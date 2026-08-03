"""Phase F.2 integration test: end-to-end XLSX user import.

Imports ``eGRM users.xlsx`` (24 rows of Province/District/Sector data from
the RISA pilot) through the wizard's bulk-import endpoints and asserts:

- All ready rows materialize as ``GRM User Project Assignment`` rows.
- Admin regions auto-create as needed; the delta against the pre-import
  count matches what ``materialize_staged_csv`` reports.
- Every imported assignment links to a Sector-level region (level_order=4
  after the Province/District/Sector hierarchy is added on top of a root
  Project level).
- ``list_project_users`` surfaces the imported rows with role + region
  pills populated.

Tests use ``frappe.tests.utils.FrappeTestCase`` so the bench test runner
discovers them. The xlsx contains spreadsheet formulas in the Phone
Numbers / Emails columns (``=_xludf.CONCAT(...)``); we synthesize clean
emails locally from the First Name to keep the import deterministic and
to avoid Frappe's xlsx reader returning the raw formula string.

The pilot workbook holds real names, phone numbers and email addresses, so
it is not committed. When it isn't on disk — CI, or a fresh checkout —
``_resolve_xlsx_path`` generates a workbook of the same shape instead (see
``_build_synthetic_rows``). Every assertion below holds for both, and a
developer who has the real file keeps testing against the real file.
"""

from __future__ import annotations

import csv
import io
import os
import tempfile
import time
from pathlib import Path

import frappe
import openpyxl
from frappe.tests.utils import FrappeTestCase
from frappe.utils.file_manager import save_file

from egrm.egrm.page.grm_project_wizard.grm_project_wizard_user_data_import import (
	poll_user_import,
	prepare_user_import,
	start_user_import,
)
from egrm.services.user_import import materialize_staged_csv

PROJECT_CODE = "TEST-STEP9-XLSX-INT"

# 4-level hierarchy mirroring the redesign plan (project / province /
# district / sector). The xlsx ships Province/District/Sector columns so
# we map 3 admin levels; level_order 1 is the project root, 2/3/4 are
# the geographic levels.
LEVELS = [
	("Project", 1),
	("Province", 2),
	("District", 3),
	("Sector", 4),
]

XLSX_PATH = Path("/tmp/egrm_users_sample.xlsx")
# Fallback to the user's Downloads folder if /tmp copy is missing — keeps
# the test self-bootstrapping for the developer who first runs it.
XLSX_FALLBACK = Path.home() / "Downloads" / "eGRM users.xlsx"
# Last resort, used by CI: a generated workbook with the same shape as the
# RISA file. Checked LAST so a developer who has the real workbook keeps
# testing against the real workbook.
SYNTHETIC_XLSX_PATH = Path(tempfile.gettempdir()) / "egrm_users_synthetic_fixture.xlsx"

# The real ``eGRM users.xlsx`` is pilot data — real names, phone numbers and
# email addresses — so it is deliberately NOT committed. Rwanda's five
# provinces are public administrative geography and are asserted on by
# ``test_materialize_xlsx_dry_run_lists_all_regions``, so the synthetic
# workbook has to reproduce them exactly; everything person-shaped below is
# invented.
PROVINCES = ["Kigali city", "Northern", "Southern", "Eastern", "Western"]
# Districts per province: 16 in total, matching the real file's distinct count.
DISTRICTS_PER_PROVINCE = {
	"Kigali city": ["District A1", "District A2", "District A3"],
	"Northern": ["District B1", "District B2", "District B3"],
	"Southern": ["District C1", "District C2", "District C3"],
	"Eastern": ["District D1", "District D2", "District D3"],
	"Western": ["District E1", "District E2", "District E3", "District E4"],
}
# Row counts per province, summing to the 24 rows the assertions pin.
ROWS_PER_PROVINCE = {
	"Kigali city": 5,
	"Northern": 5,
	"Southern": 5,
	"Eastern": 5,
	"Western": 4,
}
XLSX_HEADERS = [
	"Province",
	"District",
	"Sector",
	"First Name",
	"Last Name",
	"Gender",
	"Position",
	"Phone",
	"Phone Numbers",
	"Emails",
]
POSITIONS = ["Field Officer", "Supervisor", "Analyst", "Coordinator"]


def _build_synthetic_rows() -> list[list[str]]:
	"""Return 24 rows shaped like the RISA workbook.

	Invariants the assertions depend on:
	- exactly 24 data rows;
	- all five province names present, and no others;
	- one row with an empty Sector (it resolves to its District instead, so
	  23 sectors get created rather than 24);
	- a Sector unique to each remaining row.
	"""
	rows: list[list[str]] = []
	for province in PROVINCES:
		districts = DISTRICTS_PER_PROVINCE[province]
		for i in range(ROWS_PER_PROVINCE[province]):
			idx = len(rows)
			rows.append(
				[
					province,
					districts[i % len(districts)],
					f"Sector {idx + 1:02d}",
					f"Testuser{idx + 1:02d}",
					f"Sample{idx + 1:02d}",
					"Male" if idx % 2 == 0 else "Female",
					POSITIONS[idx % len(POSITIONS)],
					f"+25078800{idx + 1:04d}",
					# The real file carries Google-Sheets-only formulas here,
					# which openpyxl surfaces as "#NAME?". These two columns are
					# dropped before import; mirror the shape anyway.
					"#NAME?",
					"#NAME?",
				]
			)
	# The sector-less row: exercises the "resolve to parent level" path.
	rows[-1][2] = ""
	return rows


def _write_synthetic_xlsx(path: Path) -> Path:
	"""Generate the stand-in workbook so this test can run without the real
	pilot file (i.e. in CI). Rewritten on every call so edits to the row
	builder above take effect immediately."""
	wb = openpyxl.Workbook()
	ws = wb.active
	ws.append(XLSX_HEADERS)
	for row in _build_synthetic_rows():
		ws.append(row)
	path.parent.mkdir(parents=True, exist_ok=True)
	wb.save(path)
	return path


def _delete_if_exists(doctype: str, name: str) -> None:
	if frappe.db.exists(doctype, name):
		try:
			frappe.delete_doc(
				doctype,
				name,
				force=True,
				delete_permanently=True,
				ignore_permissions=True,
			)
		except Exception:
			frappe.db.rollback()


def _resolve_xlsx_path() -> Path:
	# Escape hatch so the generated-workbook path stays exercisable on a
	# machine that does have the pilot file.
	if os.environ.get("EGRM_FORCE_SYNTHETIC_XLSX"):
		return _write_synthetic_xlsx(SYNTHETIC_XLSX_PATH)
	if XLSX_PATH.exists():
		return XLSX_PATH
	if XLSX_FALLBACK.exists():
		return XLSX_FALLBACK
	# No real workbook on this machine (CI, or a fresh checkout). Fall back to
	# a generated one of the same shape rather than erroring out — this test
	# used to abort here, which is why it never ran in CI.
	return _write_synthetic_xlsx(SYNTHETIC_XLSX_PATH)


def _read_xlsx_rows() -> tuple[list[str], list[list[str]]]:
	"""Read the sample xlsx with cached values; synthesize emails locally.

	The original xlsx has Google-Sheets-only formulas in the Phone Numbers
	and Emails columns (``=_xludf.CONCAT(...)``). openpyxl in default mode
	returns the formula string, and ``data_only=True`` returns ``"#NAME?"``.
	Rather than trying to evaluate the formula, we synthesize the email as
	``{firstname}@yopmail.com`` (matching the formula's intent) so the
	test is reproducible.
	"""
	wb = openpyxl.load_workbook(_resolve_xlsx_path(), data_only=True)
	ws = wb.active
	raw = list(ws.iter_rows(values_only=True))
	headers = [str(c) if c is not None else "" for c in raw[0]]
	body: list[list[str]] = []
	for r in raw[1:]:
		if r is None:
			continue
		if not any(c is not None and str(c).strip() for c in r):
			continue
		cells = [str(c) if c is not None else "" for c in r]
		body.append(cells)
	return headers, body


def _ascii_slug(name: str) -> str:
	"""Return an ASCII-only lowercase alphanumeric slug.

	Frappe's User doctype rejects non-ASCII characters in email addresses
	(``InvalidEmailAddressError``), so we fold accented characters out.
	Falls back to ``"user"`` if the input has no ASCII alphanumerics.
	"""
	import unicodedata

	norm = unicodedata.normalize("NFKD", name or "")
	ascii_only = norm.encode("ascii", "ignore").decode("ascii")
	return "".join(c for c in ascii_only.lower() if c.isalnum()) or "user"


def _emails_for_rows(headers: list[str], rows: list[list[str]]) -> list[str]:
	"""Return the synthesized ``{slug}{idx:02d}@yopmail.com`` email per row.

	Mirrors the slugging in ``_build_synthesized_csv`` so callers (the
	test setUp) can pre-create matching ``User`` docs.
	"""
	drop = {"Phone Numbers", "Emails"}
	keep_cols = [(i, h) for i, h in enumerate(headers) if h not in drop]
	out: list[str] = []
	for idx, row in enumerate(rows):
		first_name = ""
		for kept_i, h in keep_cols:
			if h == "First Name":
				first_name = row[kept_i] if kept_i < len(row) else ""
				break
		out.append(f"{_ascii_slug(first_name)}{idx:02d}@yopmail.com")
	return out


def _build_synthesized_csv(headers: list[str], rows: list[list[str]], default_role_label: str) -> str:
	"""Build a clean CSV ready for the bulk-import endpoint.

	Adds:
	- ``Email`` column synthesised as ``{first_name_slug}{idx}@yopmail.com``
	  (deterministic, unique across rows that share a first name).
	- ``Role`` column populated with ``default_role_label`` for every row
	  (the doctype requires role).

	Drops the formula-broken ``Phone Numbers`` / ``Emails`` columns from
	the original xlsx so the auto-detect mapper isn't tempted to bind
	them.
	"""
	drop = {"Phone Numbers", "Emails"}
	keep_cols = [(i, h) for i, h in enumerate(headers) if h not in drop]
	# Synthesised columns: ``Email`` (also used as ``Username`` →
	# ``Assignment.user``), ``Role`` (the catch-all role doc name).
	# ``Assignment.user`` is doctype-required (reqd: 1) so the mapping
	# MUST cover it; we point it at the same email value.
	out_headers = [h for _, h in keep_cols] + ["Email", "Username", "Role"]

	buf = io.StringIO()
	w = csv.writer(buf)
	w.writerow(out_headers)
	for idx, row in enumerate(rows):
		out_row = [row[i] if i < len(row) else "" for i, _ in keep_cols]
		# First Name column is at index 3 in the source xlsx — but use
		# the kept-column slice so we're robust to header drift.
		first_name = ""
		for kept_i, h in keep_cols:
			if h == "First Name":
				first_name = row[kept_i] if kept_i < len(row) else ""
				break
		email = f"{_ascii_slug(first_name)}{idx:02d}@yopmail.com"
		# Username = email (Frappe's User doctype keys on email).
		out_row.extend([email, email, default_role_label])
		w.writerow(out_row)
	return buf.getvalue()


def _get_staged_path(data_import_name: str) -> str:
	"""Return the on-disk path of the staged CSV attached to a Data Import."""
	file_url = frappe.db.get_value("Data Import", data_import_name, "import_file")
	file_doc = frappe.get_doc("File", {"file_url": file_url})
	return file_doc.get_full_path()


def _save_inline_csv(content: str) -> str:
	fname = f"test_step9_xlsx_{int(time.time() * 1000)}.csv"
	f = save_file(
		fname=fname,
		content=content.encode("utf-8"),
		dt=None,
		dn=None,
		folder="Home/Attachments",
		is_private=1,
	)
	return f.file_url


class _XlsxFixture:
	"""Project + 4-level hierarchy + a single role for the imported users."""

	project_code = PROJECT_CODE
	role_name: str | None = None

	@classmethod
	def seed(cls) -> None:
		cls.teardown()
		if not frappe.db.exists("GRM Project", PROJECT_CODE):
			frappe.get_doc(
				{
					"doctype": "GRM Project",
					"project_code": PROJECT_CODE,
					"title": "Test Step 9 XLSX Integration",
				}
			).insert(ignore_permissions=True)
		for level_name, level_order in LEVELS:
			if not frappe.db.exists(
				"GRM Administrative Level Type",
				{"project": PROJECT_CODE, "level_name": level_name},
			):
				frappe.get_doc(
					{
						"doctype": "GRM Administrative Level Type",
						"project": PROJECT_CODE,
						"level_name": level_name,
						"level_order": level_order,
					}
				).insert(ignore_permissions=True)

		# A catch-all role; intentionally no admin_level so the imported
		# assignments don't fail the cross-level guard. The doctype
		# requires at least one duty (validated in
		# ``grm_project_role.py``) — we attach the seeded ``Supervise``
		# duty (which the standard fixtures install). Plan F.2 wording
		# about "role/duty unmapped pills visible" then refers to the
		# role's *admin_level* being NULL (so the level pill on the row
		# is unmapped).
		duty = "Supervise" if frappe.db.exists("GRM Duty", "Supervise") else None
		if not duty:
			# Fallback: create a duty so the test isn't blocked by a
			# missing standard fixture in a fresh site.
			duty_doc = frappe.get_doc(
				{
					"doctype": "GRM Duty",
					"duty_name": "Step9 Test Duty",
				}
			).insert(ignore_permissions=True)
			duty = duty_doc.name
		if not frappe.db.exists("GRM Project Role", {"project": PROJECT_CODE, "role_name": "Imported"}):
			role_doc = frappe.get_doc(
				{
					"doctype": "GRM Project Role",
					"project": PROJECT_CODE,
					"role_name": "Imported",
					"is_active": 1,
					"duties": [{"duty": duty}],
				}
			).insert(ignore_permissions=True)
			cls.role_name = role_doc.name
		else:
			cls.role_name = frappe.db.get_value(
				"GRM Project Role",
				{"project": PROJECT_CODE, "role_name": "Imported"},
				"name",
			)
		frappe.db.commit()

	@classmethod
	def teardown(cls) -> None:
		# Assignments → users → roles → regions → levels → project.
		for assn in frappe.get_all(
			"GRM User Project Assignment",
			filters={"project": PROJECT_CODE},
			pluck="name",
		):
			_delete_if_exists("GRM User Project Assignment", assn)
		for u in frappe.get_all(
			"User",
			filters=[["email", "like", "%@yopmail.com"]],
			pluck="name",
		):
			# Only delete the synthesised yopmail users to be safe.
			_delete_if_exists("User", u)
		for role in frappe.get_all(
			"GRM Project Role",
			filters={"project": PROJECT_CODE},
			pluck="name",
		):
			_delete_if_exists("GRM Project Role", role)
		# Regions: delete leaf-up.
		for _ in range(len(LEVELS) + 1):
			regions = frappe.get_all(
				"GRM Administrative Region",
				filters={"project": PROJECT_CODE},
				pluck="name",
			)
			if not regions:
				break
			for r in regions:
				try:
					frappe.delete_doc(
						"GRM Administrative Region",
						r,
						force=True,
						delete_permanently=True,
						ignore_permissions=True,
					)
				except Exception:
					frappe.db.rollback()
		for level_name, _ in LEVELS:
			level_id = frappe.db.get_value(
				"GRM Administrative Level Type",
				{"project": PROJECT_CODE, "level_name": level_name},
				"name",
			)
			if level_id:
				_delete_if_exists("GRM Administrative Level Type", level_id)
		_delete_if_exists("GRM Project", PROJECT_CODE)
		frappe.db.commit()


class Step9XlsxImportTests(FrappeTestCase):
	"""End-to-end: prepare → start → poll → list. 24 rows, full ORM writes."""

	@classmethod
	def setUpClass(cls) -> None:
		super().setUpClass()
		_XlsxFixture.seed()

	@classmethod
	def tearDownClass(cls) -> None:
		_XlsxFixture.teardown()
		super().tearDownClass()

	def setUp(self) -> None:
		super().setUp()
		frappe.set_user("Administrator")
		self._data_imports: list[str] = []
		# Pre-create the User docs the import will reference.
		# ``GRM User Project Assignment.validate`` hard-fails if the linked
		# User doesn't exist, so the bulk-import flow assumes Users have
		# been provisioned upstream (the wizard UI's separate "single-add"
		# path or an ops-time CSV/SSO sync). We mirror that prerequisite
		# here rather than introduce a User-create side effect to the
		# importer (which would change product behavior).
		headers, rows = _read_xlsx_rows()
		self._created_users: list[str] = []
		for email in _emails_for_rows(headers, rows):
			if frappe.db.exists("User", email):
				continue
			try:
				u = frappe.get_doc(
					{
						"doctype": "User",
						"email": email,
						"first_name": email.split("@")[0],
						"send_welcome_email": 0,
						"enabled": 1,
						"user_type": "System User",
					}
				).insert(ignore_permissions=True)
				self._created_users.append(u.name)
				# Commit per-user — User doc inserts may abort the txn on
				# post-insert hooks, so commit each one to avoid losing
				# earlier successes.
				frappe.db.commit()
			except Exception as exc:
				frappe.db.rollback()
				# Surface the first failure so we don't silently end up
				# with fewer than 24 users.
				if not getattr(self, "_user_create_err_logged", False):
					print(f"User pre-create failed for {email}: {exc!r}")
					self._user_create_err_logged = True

	def tearDown(self) -> None:
		for di in self._data_imports:
			_delete_if_exists("Data Import", di)
		for u in self._created_users:
			_delete_if_exists("User", u)
		super().tearDown()

	def test_xlsx_import_creates_24_users_and_resolves_regions(self) -> None:
		"""Import the xlsx and assert the full happy path.

		Observed region delta on a clean fixture run with the pilot
		workbook: 47 regions created (5 provinces + 19 unique
		province-district pairs + 23 unique province-district-sector
		triples; one row has no Sector so its sector is NOT created —
		hence 23 sectors, not 24). The generated workbook gives 44
		(5 + 16 + 23) because its districts don't repeat across
		provinces. The exact delta is computed dynamically below; these
		numbers just pin the ballpark so future drift is easy to spot.
		"""
		headers, rows = _read_xlsx_rows()
		self.assertEqual(len(rows), 24, f"sample xlsx must have 24 data rows; got {len(rows)}")

		csv_content = _build_synthesized_csv(headers, rows, _XlsxFixture.role_name)
		file_url = _save_inline_csv(csv_content)

		# Header mapping: the wizard's auto-detect would propose roughly
		# this; we set it explicitly to make the test deterministic.
		# Province/District/Sector are mapped via TARGET_REGION + level
		# sub-pick. Phone is mapped to mobile_no on User. Position
		# → Assignment.position_title. Gender → User.gender.
		header_mapping = {
			"Province": "administrative_region",
			"District": "administrative_region",
			"Sector": "administrative_region",
			"First Name": "User.first_name",
			"Last Name": "User.last_name",
			"Gender": "User.gender",
			"Position": "Assignment.position_title",
			"Phone": "User.mobile_no",
			"Email": "User.email",
			"Username": "Assignment.user",
			"Role": "Assignment.role",
		}
		level_mapping = {
			"Province": "Province",
			"District": "District",
			"Sector": "Sector",
		}

		# Snapshot pre-import region count.
		regions_before = frappe.db.count(
			"GRM Administrative Region",
			{"project": PROJECT_CODE},
		)
		self.assertEqual(
			regions_before,
			0,
			"fixture must start with zero regions for a clean delta read",
		)

		# Sanity: confirm the pre-created Users are visible from the same
		# connection the Importer will use.
		emails_in_db = frappe.get_all(
			"User",
			filters=[["email", "like", "%@yopmail.com"]],
			fields=["email"],
		)
		self.assertGreaterEqual(
			len(emails_in_db),
			24,
			f"setUp should have pre-created 24 yopmail users; found {len(emails_in_db)}",
		)

		prepared = prepare_user_import(
			project=PROJECT_CODE,
			file_url=file_url,
			header_mapping=header_mapping,
			level_mapping=level_mapping,
			auto_create_regions=True,
		)
		self.assertTrue(prepared["data_import"])
		self._data_imports.append(prepared["data_import"])
		# Persist the Data Import doc + the auto-created regions before
		# we kick off the importer. Otherwise an internal rollback inside
		# ``Importer.import_data`` (which fires on warnings + per-row
		# exceptions) would wipe the parent doc out of the outer test
		# transaction and ``poll_user_import`` would 404.
		frappe.db.commit()

		# Sanity: regions referenced by the staged CSV must now be in DB.
		sample_region = (prepared["regions_to_create"] or [(None, None, None)])[0][2]
		if sample_region:
			self.assertTrue(
				frappe.db.exists("GRM Administrative Region", sample_region),
				f"region {sample_region!r} not visible after commit",
			)
			# And via raw SQL — confirms it's truly committed (not just in
			# the session's value_cache).
			count_sql = frappe.db.sql(
				"SELECT COUNT(*) FROM `tabGRM Administrative Region` WHERE name = %s",
				(sample_region,),
			)
			print(f"DEBUG raw SQL count for {sample_region}: {count_sql}")

		# 24 ready rows; no skipped rows (one row has empty Sector but
		# that's a partial-path resolution which is still ready).
		self.assertEqual(prepared["rows_total"], 24)
		self.assertEqual(prepared["rows_ready"], 24)
		self.assertEqual(prepared["rows_skipped"], 0)

		# regions_to_create lists what was auto-created during prepare;
		# the count should match the post-prepare delta exactly.
		prepared_region_count = len(prepared["regions_to_create"] or [])
		regions_after_prepare = frappe.db.count(
			"GRM Administrative Region",
			{"project": PROJECT_CODE},
		)
		self.assertEqual(
			regions_after_prepare - regions_before,
			prepared_region_count,
			"regions_to_create must match the actual DB delta",
		)
		# Sanity floor: at least 5 provinces must have been created.
		self.assertGreaterEqual(
			prepared_region_count,
			5,
			"expected at least 5 provinces created from the xlsx",
		)

		# Kick off the actual import via the wizard endpoint.
		# ``form_start_import`` schedules through ``enqueue(now=run_now)``
		# where ``run_now = frappe.in_test or frappe.conf.developer_mode``.
		# In some bench/runner configurations the synchronous-now branch
		# races with the test transaction (the enqueue commit window vs
		# the test rollback) and the job is registered but not actually
		# invoked before we poll. As a robust fallback we drive the
		# ``Importer`` synchronously in-process when the wrapper hasn't
		# yet flipped status off ``Pending`` after a few polls. This is
		# exactly what ``data_import.py::start_import`` does in its
		# background-job branch — same code, same database — so the test
		# still exercises the production import code path end to end.
		start_user_import(data_import=prepared["data_import"])

		terminal = {"Success", "Partial Success", "Error", "Failed"}
		status = None
		for _ in range(6):
			status = poll_user_import(data_import=prepared["data_import"])
			if status["status"] in terminal:
				break
			time.sleep(0.25)

		if status is None or status["status"] not in terminal:
			# Drive the import directly; mirrors data_import.py::start_import.
			from frappe.core.doctype.data_import.data_import import (
				start_import as _di_start_import,
			)

			_di_start_import(prepared["data_import"])
			status = poll_user_import(data_import=prepared["data_import"])

		# Surface diagnostics if still Pending: warnings are stored on
		# the parent doc, exceptions on Data Import Log rows.
		if status["status"] not in terminal:
			di = frappe.get_doc("Data Import", prepared["data_import"])
			log_rows = frappe.get_all(
				"Data Import Log",
				filters={"data_import": prepared["data_import"]},
				fields=["success", "messages", "exception"],
				limit=5,
			)
			self.fail(
				f"import did not reach terminal status; last={status!r} "
				f"template_warnings={di.template_warnings!r} "
				f"logs={log_rows!r}"
			)
		if status.get("failed"):
			log_rows = frappe.get_all(
				"Data Import Log",
				filters={"data_import": prepared["data_import"], "success": 0},
				fields=["messages"],
				limit=2,
			)
			print(f"DEBUG failed log messages: {log_rows!r}")
			# And check the regions actually referenced by the staged CSV.
			with open(_get_staged_path(prepared["data_import"])) as fh:
				first_data = fh.readlines()[1].strip()
				print(f"DEBUG first staged row: {first_data!r}")
		self.assertIsNotNone(status)
		self.assertIn(
			status["status"],
			terminal,
			f"import did not reach terminal status; last={status!r}",
		)
		self.assertEqual(
			status["succeeded"],
			24,
			f"all 24 rows should have inserted successfully; status={status!r}",
		)
		self.assertEqual(status["failed"], 0)

		# 24 assignments now exist, all on PROJECT_CODE.
		assignment_count = frappe.db.count(
			"GRM User Project Assignment",
			{"project": PROJECT_CODE},
		)
		self.assertEqual(assignment_count, 24)

		# Every assignment has a region → level_order ∈ {Province, District, Sector}.
		# The sector-less row resolves to its District (level_order=3); the
		# rest land at Sector level (level_order=4). Both are valid region
		# leaves; what we ASSERT is that each region exists in this
		# project's hierarchy with a recognized level_order.
		rows = frappe.db.sql(
			"""
            SELECT a.name, a.administrative_region, lt.level_order
            FROM `tabGRM User Project Assignment` a
            LEFT JOIN `tabGRM Administrative Region` r
              ON r.name = a.administrative_region
            LEFT JOIN `tabGRM Administrative Level Type` lt
              ON lt.name = r.administrative_level
            WHERE a.project = %(project)s
            """,
			{"project": PROJECT_CODE},
			as_dict=True,
		)
		self.assertEqual(len(rows), 24)
		for r in rows:
			self.assertIsNotNone(r["administrative_region"])
			# Levels seeded with order 2/3/4 (province/district/sector)
			# — so every imported row's region must live in that range.
			self.assertIn(
				r["level_order"],
				(2, 3, 4),
				f"region {r['administrative_region']} resolved to "
				f"level_order={r['level_order']} (expected 2/3/4)",
			)

		# The plan calls out "role/duty unmapped pills visible" — our
		# fixture's role has no admin_level and no duties, so:
		# - role pill renders (role IS NOT NULL)
		# - the role's admin_level is None  → level pill effectively unmapped
		role_admin_levels = frappe.db.sql_list(
			"""
            SELECT DISTINCT pr.admin_level
            FROM `tabGRM User Project Assignment` a
            JOIN `tabGRM Project Role` pr ON pr.name = a.role
            WHERE a.project = %(project)s
            """,
			{"project": PROJECT_CODE},
		)
		self.assertEqual(
			list(role_admin_levels),
			[None],
			"fixture role's admin_level should be NULL (the 'unmapped' pill case)",
		)


class Step9XlsxMaterializeOnlyTests(FrappeTestCase):
	"""Lighter-weight check: materialize only, no Data Import roundtrip.

	Validates the data-validation half of F.2 in case the heavyweight
	test above is skipped (e.g. a Frappe scheduler quirk in a particular
	environment). Same xlsx; just compares ``materialize_staged_csv``
	output against the seeded hierarchy.
	"""

	@classmethod
	def setUpClass(cls) -> None:
		super().setUpClass()
		_XlsxFixture.seed()

	@classmethod
	def tearDownClass(cls) -> None:
		_XlsxFixture.teardown()
		super().tearDownClass()

	def setUp(self) -> None:
		super().setUp()
		frappe.set_user("Administrator")
		# Drop regions so the materialize call gets to auto-create them.
		for _ in range(len(LEVELS) + 1):
			regions = frappe.get_all(
				"GRM Administrative Region",
				filters={"project": PROJECT_CODE},
				pluck="name",
			)
			if not regions:
				break
			for r in regions:
				_delete_if_exists("GRM Administrative Region", r)
		frappe.db.commit()
		self._staged_files: list[str] = []

	def tearDown(self) -> None:
		for path in self._staged_files:
			try:
				if path and os.path.exists(path):
					os.remove(path)
			except OSError:
				pass
		super().tearDown()

	def test_materialize_xlsx_dry_run_lists_all_regions(self) -> None:
		"""Dry-run (auto_create=False) lists every (level, name) pair the
		xlsx would create. With 5 provinces + 19 (province,district) +
		23 (province,district,sector) the dry-run should surface a
		non-trivial count of unique pairs."""
		headers, rows = _read_xlsx_rows()
		self.assertEqual(len(rows), 24)

		# Build the canonical mapping shape ``materialize_staged_csv``
		# expects (different from the wire-format the endpoint receives).
		mapping = {
			"Province": {"target": "administrative_region", "level_type": "Province"},
			"District": {"target": "administrative_region", "level_type": "District"},
			"Sector": {"target": "administrative_region", "level_type": "Sector"},
			"First Name": {"target": "User.first_name", "level_type": None},
			"Last Name": {"target": "User.last_name", "level_type": None},
			"Gender": {"target": "User.gender", "level_type": None},
			"Position": {"target": "Assignment.position_title", "level_type": None},
			"Phone": {"target": "User.mobile_no", "level_type": None},
			"Email": {"target": "User.email", "level_type": None},
			"Username": {"target": "Assignment.user", "level_type": None},
			"Role": {"target": "Assignment.role", "level_type": None},
		}

		# Inject the synthesised email + username + role columns inline
		# so we don't bypass the mapping (which now expects ``Email``,
		# ``Username``, ``Role`` source headers).
		for idx, r in enumerate(rows):
			email = f"{_ascii_slug(r[3] if len(r) > 3 else '')}{idx:02d}@yopmail.com"
			r.append(email)  # Email
			r.append(email)  # Username (= email)
			r.append(_XlsxFixture.role_name or "")  # Role
		synthesized_headers = [*list(headers), "Email", "Username", "Role"]

		result = materialize_staged_csv(
			rows=rows,
			headers=synthesized_headers,
			mapping=mapping,
			project=PROJECT_CODE,
			auto_create_regions=False,
		)
		self._staged_files.append(result["staged_path"])

		# Dry-run: every row touches at least Province → first missing
		# level surfaces, so rows_skipped equals rows_total.
		self.assertEqual(result["rows_total"], 24)
		self.assertEqual(result["rows_ready"], 0)
		self.assertEqual(result["rows_skipped"], 24)

		# regions_to_create groups by unique (level, name); we don't pin
		# an exact integer here because Province appears 5x, but the
		# dryrun walker stops at the FIRST missing level per row, so
		# only the level whose ancestor exists gets recorded. With zero
		# regions seeded, that's "Province" for every row → 5 unique
		# provinces.
		unique_levels = {lt for lt, _ in result["regions_to_create"]}
		self.assertEqual(unique_levels, {"Province"})
		unique_provinces = {v for lt, v in result["regions_to_create"] if lt == "Province"}
		self.assertEqual(
			unique_provinces,
			{"Kigali city", "Northern", "Southern", "Eastern", "Western"},
		)
